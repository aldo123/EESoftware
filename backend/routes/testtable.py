"""
testtable.py
Excel importer for the Testing Table widget.

Excel is intentionally limited to:
  1. Testing Item
  2. Lower Limit
  3. Upper Limit

Mode / Source / Device / Address / Address Type are runtime configuration
and are kept in the Page Builder Testing Table Settings.
"""
from flask import Blueprint, request, jsonify
from openpyxl import load_workbook

testtable_bp = Blueprint("testtable", __name__)


@testtable_bp.post("/api/testtable/parse-excel")
def parse_excel():
    file = request.files.get("file")
    if not file:
        return jsonify({"success": False, "message": "No file uploaded"}), 400

    try:
        wb = load_workbook(file, data_only=True, read_only=True)
        ws = wb.active

        first = next(ws.iter_rows(values_only=True), None)
        if not first:
            return jsonify({"success": False, "message": "Excel is empty"}), 400

        headers = [str(v or "").strip().lower() for v in first]
        expected = ["testing item", "lower limit", "upper limit"]

        # Require the three columns in order. Extra columns are rejected so
        # Excel remains a clean specification/master file.
        if headers[:3] != expected or len([h for h in headers if h]) != 3:
            return jsonify({
                "success": False,
                "message": "Excel must contain exactly 3 columns: Testing Item, Lower Limit, Upper Limit"
            }), 400

        rows = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not values or not any(v not in (None, "") for v in values[:3]):
                continue

            item = str(values[0] or "").strip()
            if not item:
                continue

            rows.append({
                "id": f"row_{len(rows) + 1}",
                "item": item,
                "lower": "" if values[1] is None else str(values[1]).strip(),
                "upper": "" if values[2] is None else str(values[2]).strip(),
            })

        return jsonify({"success": True, "rows": rows, "count": len(rows)})
    except Exception as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
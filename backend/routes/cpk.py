"""
routes/cpk.py — CPK Analyzer backend: process capability calculation, an
Excel importer that hands the whole table to the frontend (filtering,
summary, and failure analysis all run client-side from that table), a
Traceability MySQL puller for the database data source, and two Excel
exporters for the result/report and the raw dataset.
"""
import io
import re

from flask import Blueprint, request, jsonify, send_file
from openpyxl import load_workbook, Workbook

from cpk_engine import (
    CPKCalculator,
    CPKAssessment,
    CalibrationEngine,
    UnsafeDataError,
    clean_values,
    histogram,
    boxplot_stats,
)

cpk_bp = Blueprint("cpk", __name__, url_prefix="/api/cpk")

# Table/column names can't be parameterized like values can, so they're
# checked against this whitelist before ever touching an f-string SQL query.
_IDENTIFIER_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")


def _safe_identifier(name, label):
    name = str(name or "").strip()
    if not _IDENTIFIER_RE.match(name):
        raise ValueError(f"{label} tidak valid: '{name}'")
    return name


@cpk_bp.get("/industries")
def get_industries():
    return jsonify({"success": True, "industries": CPKAssessment.get_industry_names()})


@cpk_bp.get("/standards")
def get_standards():
    return jsonify({"success": True, "standards": CPKAssessment.STANDARDS})


def _cell_to_json(value):
    """openpyxl hands back datetime/Decimal-ish objects for some cells —
    stringify anything that isn't already JSON-safe."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


@cpk_bp.post("/parse-file")
def parse_file():
    """Read an uploaded Excel file and hand back the whole table (headers +
    every row as an object). The page keeps this in memory and does all its
    filtering/summary/failure-analysis client-side from it, same as the
    desktop tool kept the loaded DataFrame in memory."""
    file = request.files.get("file")
    if not file:
        return jsonify({"success": False, "message": "No file uploaded"}), 400

    try:
        wb = load_workbook(file, data_only=True, read_only=True)
        ws = wb.active

        first = next(ws.iter_rows(values_only=True), None)
        if not first:
            return jsonify({"success": False, "message": "File is empty"}), 400

        headers = [str(v or "").strip() for v in first]
        if not any(headers):
            return jsonify({"success": False, "message": "File has no column headers"}), 400

        rows = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not values or not any(v not in (None, "") for v in values):
                continue
            row = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                row[h] = _cell_to_json(values[i] if i < len(values) else None)
            rows.append(row)

        return jsonify({
            "success": True,
            "columns": [h for h in headers if h],
            "rows": rows,
            "count": len(rows),
        })
    except Exception as exc:
        return jsonify({"success": False, "message": str(exc)}), 500


@cpk_bp.post("/analyze")
def analyze():
    body = request.get_json(silent=True) or {}

    values = clean_values(body.get("values"))
    industry = body.get("industry") or "General Manufacturing"

    try:
        lsl = float(body.get("lsl"))
        usl = float(body.get("usl"))
    except (TypeError, ValueError):
        return jsonify({"success": False, "message": "LSL/USL harus berupa angka"}), 400

    try:
        stats = CPKCalculator.calculate(values, lsl, usl)
        assessment = CPKAssessment.assess(
            stats["cpk"], stats["cp"], stats["cpl"], stats["cpu"],
            lsl, usl, stats["mean"], stats["std"], industry,
        )
        hist = histogram(values, lsl, usl)
        boxplot = boxplot_stats(values)
    except UnsafeDataError as e:
        return jsonify({"success": False, "message": str(e)}), 400

    return jsonify({
        "success": True,
        "stats": stats,
        "assessment": assessment,
        "histogram": hist,
        "boxplot": boxplot,
        "industry": industry,
    })


@cpk_bp.post("/calibrate")
def calibrate():
    body = request.get_json(silent=True) or {}
    values = clean_values(body.get("values"))

    try:
        lsl = float(body.get("lsl"))
        usl = float(body.get("usl"))
    except (TypeError, ValueError):
        return jsonify({"success": False, "message": "LSL/USL harus berupa angka"}), 400

    try:
        result = CalibrationEngine.auto_calibrate(values, lsl, usl)
        calibrated = clean_values(result["calibrated_values"])
        result["calibrated_histogram"] = histogram(calibrated, lsl, usl)
        result["calibrated_boxplot"] = boxplot_stats(calibrated)
    except UnsafeDataError as e:
        return jsonify({"success": False, "message": str(e)}), 400

    return jsonify({"success": True, **result})


@cpk_bp.post("/query-database")
def query_database():
    """Pull one numeric column's values out of the Traceability MySQL
    database for CPK analysis, with an optional station/line/date-range
    filter. Table and column names come from the widget's own config
    (not free user input at runtime), but are still validated as plain
    identifiers since they can't be bound as query parameters."""
    from db_manager import db

    body = request.get_json(silent=True) or {}

    try:
        table = _safe_identifier(body.get("table"), "Table name")
        column = _safe_identifier(body.get("column"), "Column name")
    except ValueError as e:
        return jsonify({"success": False, "message": str(e)}), 400

    where_clauses = [f"`{column}` IS NOT NULL"]
    params = []

    station = body.get("station")
    if station:
        where_clauses.append("`STATION` = %s")
        params.append(station)

    line = body.get("line")
    if line:
        where_clauses.append("`LINE` = %s")
        params.append(line)

    date_from = body.get("date_from")
    if date_from:
        where_clauses.append("`DATETIME` >= %s")
        params.append(date_from)

    date_to = body.get("date_to")
    if date_to:
        where_clauses.append("`DATETIME` <= %s")
        params.append(date_to)

    limit = body.get("limit")
    try:
        limit = min(int(limit), 50000) if limit else 5000
    except (TypeError, ValueError):
        limit = 5000

    sql = (
        f"SELECT `{column}` AS value FROM `{table}` "
        f"WHERE {' AND '.join(where_clauses)} "
        f"ORDER BY `DATETIME` DESC LIMIT %s"
    )
    params.append(limit)

    rows = db.fetch_all(sql, tuple(params))
    values = [row.get("value") for row in rows]

    return jsonify({"success": True, "values": values, "count": len(values)})


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 60)


def _xlsx_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@cpk_bp.post("/export-result")
def export_result():
    """CPK_RESULT + FILTERED_DATA + RECOMMENDATIONS sheets, mirroring the
    desktop tool's Export Result dialog."""
    body = request.get_json(silent=True) or {}
    stats = body.get("stats") or {}
    assessment = body.get("assessment") or {}
    rows = body.get("rows") or []
    parameter = body.get("parameter") or ""
    industry = body.get("industry") or ""
    filters = body.get("filters") or {}
    calibration = body.get("calibration") or {}

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "CPK_RESULT"

    fields = [
        ("Parameter", parameter),
        ("Sample Size", stats.get("sample_size")),
        ("Mean", stats.get("mean")),
        ("Std Dev", stats.get("std")),
        ("Minimum", stats.get("minimum")),
        ("Maximum", stats.get("maximum")),
        ("Range", stats.get("range")),
        ("LSL", stats.get("lsl")),
        ("USL", stats.get("usl")),
        ("Tolerance", stats.get("tolerance")),
        ("Cp", stats.get("cp")),
        ("Cpu", stats.get("cpu")),
        ("Cpl", stats.get("cpl")),
        ("Cpk", stats.get("cpk")),
        ("Z-Score", stats.get("z_score")),
        ("PPM", stats.get("ppm")),
        ("Below LSL", stats.get("below_lsl")),
        ("Above USL", stats.get("above_usl")),
        ("Out of Spec", stats.get("out_of_spec")),
        ("Capability", stats.get("capability")),
        ("LINE Filter", filters.get("line") or "ALL"),
        ("STATION Filter", filters.get("station") or "ALL"),
        ("RESULT Filter", filters.get("result") or "ALL"),
        ("Industry Standard", industry),
        ("Assessment Status", assessment.get("status")),
        ("Calibration k", calibration.get("k", "N/A")),
        ("Calibration b", calibration.get("b", "N/A")),
        ("Calibration Applied", "Yes" if calibration else "No"),
    ]
    ws1.append(["Field", "Value"])
    for label, value in fields:
        ws1.append([label, _cell_to_json(value)])
    _autosize(ws1)

    ws2 = wb.create_sheet("FILTERED_DATA")
    if rows:
        headers = list(rows[0].keys())
        ws2.append(headers)
        for row in rows:
            ws2.append([_cell_to_json(row.get(h)) for h in headers])
        _autosize(ws2)

    ws3 = wb.create_sheet("RECOMMENDATIONS")
    ws3.append(["Issue", "Detail", "Action"])
    for rec in assessment.get("recommendations", []):
        ws3.append([rec.get("issue"), rec.get("detail"), rec.get("action")])
    _autosize(ws3)

    return _xlsx_response(wb, "cpk_result.xlsx")


@cpk_bp.post("/export-all")
def export_all():
    """RAW_DATA + SUMMARY sheets, mirroring the desktop tool's Export All
    Data dialog."""
    body = request.get_json(silent=True) or {}
    rows = body.get("rows") or []
    summary = body.get("summary") or {}

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "RAW_DATA"
    if rows:
        headers = list(rows[0].keys())
        ws1.append(headers)
        for row in rows:
            ws1.append([_cell_to_json(row.get(h)) for h in headers])
        _autosize(ws1)

    ws2 = wb.create_sheet("SUMMARY")
    ws2.append(["Field", "Value"])
    for key, value in summary.items():
        if isinstance(value, dict):
            for sub_key, sub_value in value.items():
                ws2.append([f"{key}.{sub_key}", _cell_to_json(sub_value)])
        else:
            ws2.append([key, _cell_to_json(value)])
    _autosize(ws2)

    return _xlsx_response(wb, "cpk_all_data.xlsx")
